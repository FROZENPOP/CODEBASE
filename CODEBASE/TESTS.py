from openpyxl import load_workbook
import azimuth.model_comparison
import numpy as np

row_num = 2
workbook = load_workbook("C:\Users\chann\Downloads\FINAL_FILTERED_CRISPR_DATA.xlsx")
sheet = workbook.get_sheet_by_name("FC_plus_RES_withPredictions")

for i in range(5309):
    seq30mer = sheet.cell(row=row_num, column=2).value
    aacp = int(sheet.cell(row=row_num, column=5).value)
    pp = int(sheet.cell(row=row_num, column=4).value)

    sequences = np.array([str(seq30mer)])
    amino_acid_cut_positions = np.array([aacp])
    percent_peptides = np.array([pp])
    predictions = azimuth.model_comparison.predict(sequences)
    writein = "K" + str(row_num)

    sheet[writein].value = str(predictions)
    row_num +=1


workbook.save("C:\Users\chann\Downloads\FINAL_FILTERED_CRISPR_DATA.xlsx")
